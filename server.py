import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Creating a sample dataframe with additional articles
data = {
    "ID": ["seoul", "busan", "daegu"],
    "Name": ["Seoul", "Busan", "Daegu"],
    "Move1": ["Seoul Moving Center: 010-1234-5678", "Busan Moving Center: 010-2345-6789", "Daegu Moving Center: 010-3456-7890"],
    "Move2": ["Another article about Seoul", "Another article about Busan", "Another article about Daegu"],
    "Move3": ["Yet another article about Seoul", "Yet another article about Busan", "Yet another article about Daegu"],
    "Move4": ["Additional article about Seoul", "Additional article about Busan", "Additional article about Daegu"],
    "Move5": ["More information about Seoul", "More information about Busan", "More information about Daegu"]
}

# Creating a Pandas dataframe
df = pd.DataFrame(data)

# Creating an Excel writer object and saving the dataframe
file_path = "articles_example_with_colors_and_multiple_articles.xlsx"
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    # Accessing the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Defining fill colors
    fill_colors = [PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                   PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid"),
                   PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),
                   PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
                   PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")]
    
    # Applying fill colors to specific cells
    for row in range(2, len(df) + 2):
        worksheet[f"C{row}"].fill = fill_colors[0]
        worksheet[f"D{row}"].fill = fill_colors[1]
        worksheet[f"E{row}"].fill = fill_colors[2]
        worksheet[f"F{row}"].fill = fill_colors[3]
        worksheet[f"G{row}"].fill = fill_colors[4]

# Save the file
workbook.save(file_path)

file_path
